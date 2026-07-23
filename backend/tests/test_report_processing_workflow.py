from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import json
import unittest
from unittest.mock import patch

from openpyxl import Workbook

from app.db.connection import connect
from app.db.schema import init_db
from app.ingest.excel_importer import import_excel, parse_excel
from app.ingest.repository import ReportImportRepository
from app.validation.profile_repository import ValidationProfileRepository
from app.validation.profiles import HeuristicProfileDraftProvider
from app.validation.repository import ValidationRepository
from app.validation.workflow import ValidationWorkflow
from app.workflows.report_processing import ReportProcessingError, ReportProcessingWorkflow


class CountingProfileProvider(HeuristicProfileDraftProvider):
    def __init__(self) -> None:
        self.calls = 0

    def draft(self, *args, **kwargs):  # type: ignore[no-untyped-def]
        self.calls += 1
        return super().draft(*args, **kwargs)


class ReportProcessingWorkflowTest(unittest.TestCase):
    def make_workbook(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "1-1-1-1 테스트"
        sheet.append(["표번호", "1-1-1-1"])
        sheet.append(["통계명", "성별 테스트"])
        sheet.append(["단위", "명"])
        sheet.append(["기준일", "2025. 12. 31."])
        sheet.append(["출처", "테스트과"])
        sheet.append([])
        sheet.append(["구분", "합계", "남", "여"])
        sheet.append(["계", 3, 1, 2])
        workbook.save(path)
        workbook.close()

    def test_repository_rejects_an_import_without_tables(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            source_path = root / "empty.xlsx"
            source_path.write_bytes(b"empty")

            with self.assertRaisesRegex(ValueError, "통계표"):
                ReportImportRepository(root / "report.sqlite").replace_report(
                    source_path=source_path,
                    year=2027,
                    title="빈 연보",
                    tables=[],
                )

    def test_excel_import_and_validation_are_pinned_to_the_imported_report(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook_path = root / "2026.xlsx"
            db_path = root / "report.sqlite"
            self.make_workbook(workbook_path)

            imported = import_excel(
                workbook_path,
                db_path=db_path,
                year=2026,
                run_validation=False,
            )
            result = ValidationWorkflow(db_path).run(report_id=int(imported["report_id"]))

            self.assertEqual(result.report_id, imported["report_id"])
            self.assertEqual(result.tables, 1)
            self.assertGreater(result.run_id, 0)

    def test_excel_parser_preserves_merged_header_values(self) -> None:
        with TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "merged.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "1-1-1-1"
            sheet.append(["표번호", "1-1-1-1"])
            sheet.append(["통계명", "병합 테스트"])
            sheet.append(["구분", "인원", ""])
            sheet.append(["지역", "남", "여"])
            sheet.append(["계", 1, 2])
            sheet.merge_cells("B3:C3")
            workbook.save(workbook_path)
            workbook.close()

            tables = parse_excel(workbook_path)

            self.assertEqual(len(tables), 1)
            self.assertEqual(tables[0].matrix[0], ["구분", "인원", "인원"])

    def test_excel_metadata_number_row_is_not_used_as_title(self) -> None:
        with TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "metadata.xlsx"
            self.make_workbook(workbook_path)

            tables = parse_excel(workbook_path)

            self.assertEqual(tables[0].title, "성별 테스트")

    def test_excel_percentage_format_is_stored_as_displayed_percent(self) -> None:
        with TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "percentage.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "1-1-1-2 비율"
            sheet.append(["구분", "비율"])
            sheet.append(["계", 0.125])
            sheet["B2"].number_format = "0.0%"
            workbook.save(workbook_path)
            workbook.close()

            tables = parse_excel(workbook_path)

            self.assertEqual(tables[0].code, "1-1-1-2")
            self.assertEqual(tables[0].matrix[1][1], "12.5%")

    def test_matching_ready_profile_is_reused_without_regeneration(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook_path = root / "reuse.xlsx"
            db_path = root / "report.sqlite"
            self.make_workbook(workbook_path)
            imported = import_excel(
                workbook_path,
                db_path=db_path,
                year=2026,
                run_validation=False,
            )
            report_id = int(imported["report_id"])
            tables = ValidationRepository(db_path).load_tables(report_id)
            repository = ValidationProfileRepository(db_path)
            provider = CountingProfileProvider()

            repository.ensure_profiles(report_id=report_id, tables=tables, provider=provider)
            repository.ensure_profiles(report_id=report_id, tables=tables, provider=provider)

            self.assertEqual(provider.calls, 1)

    def test_explicit_profile_refresh_creates_a_new_draft(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook_path = root / "refresh.xlsx"
            db_path = root / "report.sqlite"
            self.make_workbook(workbook_path)
            imported = import_excel(
                workbook_path,
                db_path=db_path,
                year=2026,
                run_validation=False,
            )
            report_id = int(imported["report_id"])
            tables = ValidationRepository(db_path).load_tables(report_id)
            repository = ValidationProfileRepository(db_path)
            provider = CountingProfileProvider()

            repository.ensure_profiles(report_id=report_id, tables=tables, provider=provider)
            repository.ensure_profiles(
                report_id=report_id,
                tables=tables,
                provider=provider,
                refresh=True,
            )

            self.assertEqual(provider.calls, 2)
            self.assertEqual(
                len(
                    [
                        profile
                        for profile in repository.list_profiles()
                        if profile.table_code == "1-1-1-1"
                    ]
                ),
                1,
            )

    def test_row_structure_change_creates_a_new_profile(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook_path = root / "changed.xlsx"
            db_path = root / "report.sqlite"
            self.make_workbook(workbook_path)
            repository = ValidationProfileRepository(db_path)
            provider = CountingProfileProvider()

            first = import_excel(workbook_path, db_path=db_path, year=2026, run_validation=False)
            first_tables = ValidationRepository(db_path).load_tables(int(first["report_id"]))
            repository.ensure_profiles(
                report_id=int(first["report_id"]),
                tables=first_tables,
                provider=provider,
            )

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "1-1-1-1 테스트"
            sheet.append(["표번호", "1-1-1-1"])
            sheet.append(["통계명", "성별 테스트"])
            sheet.append(["구분", "합계", "남", "여"])
            sheet.append(["전체", 3, 1, 2])
            workbook.save(workbook_path)
            workbook.close()
            second = import_excel(workbook_path, db_path=db_path, year=2027, run_validation=False)
            second_tables = ValidationRepository(db_path).load_tables(int(second["report_id"]))
            repository.ensure_profiles(
                report_id=int(second["report_id"]),
                tables=second_tables,
                provider=provider,
            )

            self.assertEqual(provider.calls, 2)

    def test_same_title_and_structure_inherits_profile_after_code_change(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            first_path = root / "first.xlsx"
            second_path = root / "second.xlsx"
            db_path = root / "report.sqlite"
            self.make_workbook(first_path)
            provider = CountingProfileProvider()
            repository = ValidationProfileRepository(db_path)

            first = import_excel(first_path, db_path=db_path, year=2026, run_validation=False)
            first_tables = ValidationRepository(db_path).load_tables(int(first["report_id"]))
            repository.ensure_profiles(
                report_id=int(first["report_id"]),
                tables=first_tables,
                provider=provider,
            )

            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "2-1-1-1 테스트"
            sheet.append(["표번호", "2-1-1-1"])
            sheet.append(["통계명", "성별 테스트"])
            sheet.append(["단위", "명"])
            sheet.append(["기준일", "2026. 12. 31."])
            sheet.append(["출처", "테스트과"])
            sheet.append([])
            sheet.append(["구분", "합계", "남", "여"])
            sheet.append(["계", 4, 2, 2])
            workbook.save(second_path)
            workbook.close()

            second = import_excel(second_path, db_path=db_path, year=2027, run_validation=False)
            second_tables = ValidationRepository(db_path).load_tables(int(second["report_id"]))
            profiles = repository.ensure_profiles(
                report_id=int(second["report_id"]),
                tables=second_tables,
                provider=provider,
            )

            inherited = profiles["2-1-1-1"]
            self.assertEqual(provider.calls, 1)
            self.assertEqual(inherited.source, "inherited")
            inherited_ids = [str(spec.get("id", "")) for spec in inherited.check_specs]
            self.assertFalse(any(rule_id.startswith("profile.1-1-1-1.") for rule_id in inherited_ids))

    def test_processing_job_records_import_and_validation_stages(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook_path = root / "job.xlsx"
            db_path = root / "report.sqlite"
            self.make_workbook(workbook_path)
            workflow = ReportProcessingWorkflow(db_path)
            job_id = workflow.create_job(
                source_path=workbook_path,
                year=2026,
                title="2026 테스트 연보",
            )

            job = workflow.run(job_id)

            self.assertEqual(job["status"], "completed")
            self.assertEqual(
                [stage["stage_name"] for stage in job["stages"]],
                ["import", "calculation_validation", "language_validation"],
            )
            self.assertTrue(all(stage["status"] == "completed" for stage in job["stages"]))
            language_stage = job["stages"][-1]
            language_result = json.loads(str(language_stage["result_json"]))
            self.assertIn(language_result["status"], {"완료", "대기"})

    def test_retry_resumes_failed_stage_without_reimporting(self) -> None:
        with TemporaryDirectory() as directory:
            root = Path(directory)
            workbook_path = root / "retry.xlsx"
            db_path = root / "report.sqlite"
            self.make_workbook(workbook_path)
            workflow = ReportProcessingWorkflow(db_path)
            job_id = workflow.create_job(
                source_path=workbook_path,
                year=2026,
                title="2026 재시도 테스트 연보",
            )

            with patch.object(
                workflow,
                "_run_language_validation",
                side_effect=RuntimeError("temporary language failure"),
            ):
                with self.assertRaises(ReportProcessingError):
                    workflow.run(job_id)

            workflow.queue_retry(job_id)
            completed = workflow.retry(job_id)
            stage_names = [stage["stage_name"] for stage in completed["stages"]]

            self.assertEqual(completed["status"], "completed")
            self.assertEqual(stage_names.count("import"), 1)
            self.assertEqual(stage_names.count("calculation_validation"), 1)
            self.assertEqual(stage_names.count("language_validation"), 2)


if __name__ == "__main__":
    unittest.main()
