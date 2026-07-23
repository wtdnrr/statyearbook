from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
import argparse
import re
import sqlite3
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.contact_metadata import ContactMetadata, parse_contact_metadata
from app.db.connection import DB_PATH, connect
from app.db.schema import init_db
from app.validation.highlights import highlight_cells_for
from app.validation.spec_repository import load_rule_specs_by_id
from app.validation.models import ValidationTable, restore_hyphenated_line_breaks
from app.validation.rules import combined_row_label
from app.validation.repository import ValidationRepository


EXCLUDED_CHECK_TYPES = {"이상치 검수"}
PART_SUFFIX_RE = re.compile(r"\s+표\d+$")
INVALID_SHEET_CHARS_RE = re.compile(r"[\\/*?:\[\]]")
NAVY = "17324D"
BLUE = "DCEAF7"
ERROR_TARGET = "F4CCCC"
ERROR_RELATED = "FDE9E7"
REVIEW_TARGET = "FCE5CD"
REVIEW_RELATED = "FFF4E5"
WHITE = "FFFFFF"
GRAY = "F3F5F8"
TEXT = "172033"
MUTED = "5F6F86"
THIN_SIDE = Side(style="thin", color="D6DFEA")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


@dataclass(frozen=True)
class ExportResult:
    report_id: int
    run_id: int
    metadata_rows: int
    validation_rows: int
    validation_sheets: int
    metadata_path: Path
    highlighted_path: Path
    validation_index_path: Path


def export_validation_workbooks(
    db_path: Path = DB_PATH,
    *,
    year: int | None = None,
    report_id: int | None = None,
    run_id: int | None = None,
    output_dir: Path = Path("exports"),
) -> ExportResult:
    output_dir.mkdir(parents=True, exist_ok=True)
    with connect(db_path) as connection:
        init_db(connection)
        report = resolve_report(connection, report_id=report_id, year=year)
        resolved_run_id = resolve_run_id(connection, report_id=int(report["id"]), run_id=run_id)
        tables = connection.execute(
            "SELECT * FROM stat_tables WHERE report_id = ? ORDER BY table_order, id",
            (report["id"],),
        ).fetchall()
        failing_checks = connection.execute(
            """
            SELECT vc.*, st.code, st.title, st.table_order
            FROM validation_checks vc
            JOIN stat_tables st ON st.id = vc.table_id
            WHERE vc.run_id = ?
              AND vc.status != '정상'
              AND vc.check_type NOT IN ({})
            ORDER BY st.table_order,
                     CASE vc.status WHEN '오류 의심' THEN 0 ELSE 1 END,
                     vc.row_index IS NULL, vc.row_index, vc.col_index, vc.id
            """.format(",".join("?" for _ in EXCLUDED_CHECK_TYPES)),
            (resolved_run_id, *sorted(EXCLUDED_CHECK_TYPES)),
        ).fetchall()

        contacts = {int(table["id"]): parse_contact_metadata(table["source"] or "") for table in tables}
        year_label = str(report["year"])
        metadata_path = output_dir / f"{year_label}_통계연보_메타데이터.xlsx"
        highlighted_path = output_dir / f"{year_label}_통계연보_검수표_하이라이트.xlsx"
        validation_index_path = output_dir / f"{year_label}_통계연보_검수내역.xlsx"

        metadata_rows = len(group_tables_by_base(tables))
        write_metadata_workbook(metadata_path, tables, contacts)
        validation_sheet_names = write_highlighted_workbook(
            highlighted_path,
            connection=connection,
            report_id=int(report["id"]),
            run_id=resolved_run_id,
            tables=tables,
            checks=failing_checks,
            contacts=contacts,
        )
        write_validation_index_workbook(
            validation_index_path,
            checks=failing_checks,
            contacts=contacts,
            highlighted_filename=highlighted_path.name,
            sheet_names=validation_sheet_names,
            validation_tables=validation_table_map(db_path, int(report["id"])),
        )

    validate_generated_workbooks(
        metadata_path,
        highlighted_path,
        validation_index_path,
        expected_metadata_rows=metadata_rows,
        expected_validation_rows=len(failing_checks),
        expected_validation_sheets=len(validation_sheet_names),
    )
    return ExportResult(
        report_id=int(report["id"]),
        run_id=resolved_run_id,
        metadata_rows=metadata_rows,
        validation_rows=len(failing_checks),
        validation_sheets=len(validation_sheet_names),
        metadata_path=metadata_path.resolve(),
        highlighted_path=highlighted_path.resolve(),
        validation_index_path=validation_index_path.resolve(),
    )


def resolve_report(
    connection: sqlite3.Connection,
    *,
    report_id: int | None,
    year: int | None,
) -> sqlite3.Row:
    if report_id is not None:
        row = connection.execute("SELECT * FROM annual_reports WHERE id = ?", (report_id,)).fetchone()
    elif year is not None:
        row = connection.execute(
            "SELECT * FROM annual_reports WHERE year = ? ORDER BY imported_at DESC, id DESC LIMIT 1",
            (year,),
        ).fetchone()
    else:
        row = connection.execute(
            "SELECT * FROM annual_reports ORDER BY imported_at DESC, id DESC LIMIT 1"
        ).fetchone()
    if row is None:
        raise RuntimeError("내보낼 통계연보를 찾을 수 없습니다.")
    return row


def resolve_run_id(connection: sqlite3.Connection, *, report_id: int, run_id: int | None) -> int:
    if run_id is not None:
        row = connection.execute(
            "SELECT id FROM validation_runs WHERE id = ? AND report_id = ?",
            (run_id, report_id),
        ).fetchone()
    else:
        row = connection.execute(
            """
            SELECT id FROM validation_runs
            WHERE report_id = ?
            ORDER BY completed_at DESC, id DESC
            LIMIT 1
            """,
            (report_id,),
        ).fetchone()
    if row is None:
        raise RuntimeError("내보낼 검수 실행을 찾을 수 없습니다.")
    return int(row["id"])


def unique(values: Iterable[str]) -> list[str]:
    return list(dict.fromkeys(value.strip() for value in values if value.strip()))


def write_metadata_workbook(
    path: Path,
    tables: list[sqlite3.Row],
    contacts: dict[int, ContactMetadata],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "메타데이터"
    headers = ["번호", "통계명", "기준일", "단위", "주석", "소속", "이름", "내선번호", "출처"]
    write_header_row(sheet, 1, headers)
    for row_index, (base_code, parts) in enumerate(group_tables_by_base(tables).items(), start=2):
        part_contacts = [contacts[int(part["id"])] for part in parts]
        values = [
            base_code,
            parts[0]["title"],
            join_distinct(part["base_date"] for part in parts),
            join_distinct(part["unit"] for part in parts),
            join_distinct((clean_multiline_text(part["note"] or "") for part in parts), separator="\n\n"),
            join_distinct(contact.department for contact in part_contacts),
            join_distinct(contact.officer for contact in part_contacts),
            join_distinct(contact.extension for contact in part_contacts),
            join_distinct((contact.source_reference for contact in part_contacts), separator="\n"),
        ]
        write_data_row(sheet, row_index, values)
    format_tabular_sheet(sheet, widths=[18, 38, 17, 16, 70, 24, 18, 20, 48])
    workbook.save(path)


def write_highlighted_workbook(
    path: Path,
    *,
    connection: sqlite3.Connection,
    report_id: int,
    run_id: int,
    tables: list[sqlite3.Row],
    checks: list[sqlite3.Row],
    contacts: dict[int, ContactMetadata],
) -> dict[str, str]:
    del run_id
    workbook = Workbook()
    workbook.remove(workbook.active)
    checks_by_table: dict[int, list[sqlite3.Row]] = defaultdict(list)
    issue_base_codes: set[str] = set()
    for check in checks:
        checks_by_table[int(check["table_id"])].append(check)
        issue_base_codes.add(base_table_code(check["code"]))

    tables_by_base: dict[str, list[sqlite3.Row]] = defaultdict(list)
    for table in tables:
        tables_by_base[base_table_code(table["code"])].append(table)

    sheet_names: dict[str, str] = {}
    used_names: set[str] = set()
    for base_code in sorted(issue_base_codes, key=lambda code: min(row["table_order"] for row in tables_by_base[code])):
        sheet_name = unique_sheet_name(base_code, used_names)
        sheet_names[base_code] = sheet_name
        sheet = workbook.create_sheet(sheet_name)
        write_statistic_sheet(
            sheet,
            connection=connection,
            tables=tables_by_base[base_code],
            checks_by_table=checks_by_table,
            contacts=contacts,
        )

    if not workbook.worksheets:
        workbook.create_sheet("검수 결과 없음")
    workbook.save(path)
    return sheet_names


def write_statistic_sheet(
    sheet,
    *,
    connection: sqlite3.Connection,
    tables: list[sqlite3.Row],
    checks_by_table: dict[int, list[sqlite3.Row]],
    contacts: dict[int, ContactMetadata],
) -> None:
    first = tables[0]
    base_code = base_table_code(first["code"])
    max_columns = max(table_max_col(connection, int(table["id"])) + 1 for table in tables)
    display_columns = max(max_columns, 9)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=display_columns)
    sheet.cell(1, 1, f"{base_code}  {first['title']}")
    sheet.cell(1, 1).font = Font(name="Arial", size=16, bold=True, color=WHITE)
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    sheet.cell(1, 1).alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 30

    contact = contacts[int(first["id"])]
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=display_columns)
    sheet.cell(
        2,
        1,
        f"기준일 {first['base_date'] or '-'}   |   단위 {first['unit'] or '-'}   |   "
        f"소속 {contact.department or '-'}   |   이름 {contact.officer or '-'}   |   내선번호 {contact.extension or '-'}",
    )
    sheet.cell(2, 1).fill = PatternFill("solid", fgColor=GRAY)
    sheet.cell(2, 1).font = Font(name="Arial", size=10, color=TEXT)
    sheet.cell(2, 1).alignment = Alignment(vertical="center")
    metadata_checks = [
        check
        for table in tables
        for check in checks_by_table.get(int(table["id"]), [])
        if check["check_type"] == "메타정보 검수"
    ]
    if metadata_checks:
        metadata_status = (
            "오류 의심"
            if any(check["status"] == "오류 의심" for check in metadata_checks)
            else "확인 필요"
        )
        apply_highlight(sheet.cell(2, 1), status=metadata_status, role="target")
        for check in metadata_checks:
            append_issue_comment(sheet.cell(2, 1), check)

    sheet.cell(3, 1, "■ 오류/확인 대상 셀")
    sheet.cell(3, 2, "■ 계산 관련 셀")
    sheet.cell(3, 1).fill = PatternFill("solid", fgColor=ERROR_TARGET)
    sheet.cell(3, 2).fill = PatternFill("solid", fgColor=REVIEW_RELATED)
    sheet.cell(3, 1).font = sheet.cell(3, 2).font = Font(name="Arial", size=9, color=TEXT)

    current_row = 5
    for table in tables:
        table_id = int(table["id"])
        table_checks = checks_by_table.get(table_id, [])
        current_row = write_table_section(
            sheet,
            connection=connection,
            table=table,
            checks=table_checks,
            start_row=current_row,
            display_columns=display_columns,
        )
        current_row += 2

    for column_index in range(1, display_columns + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 20 if column_index > 1 else 28
    sheet.freeze_panes = "A5"
    sheet.sheet_view.showGridLines = False
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.auto_filter.ref = None


def write_table_section(
    sheet,
    *,
    connection: sqlite3.Connection,
    table: sqlite3.Row,
    checks: list[sqlite3.Row],
    start_row: int,
    display_columns: int,
) -> int:
    table_id = int(table["id"])
    cells = connection.execute(
        """
        SELECT row_index, col_index, text_value, is_header, footnote_marker
        FROM stat_table_cells WHERE table_id = ?
        ORDER BY row_index, col_index
        """,
        (table_id,),
    ).fetchall()
    max_row = max((int(cell["row_index"]) for cell in cells), default=0)
    max_col = max((int(cell["col_index"]) for cell in cells), default=0)

    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=display_columns)
    sheet.cell(start_row, 1, f"{table['code']}  {table['title']}")
    sheet.cell(start_row, 1).font = Font(name="Arial", size=12, bold=True, color=TEXT)
    sheet.cell(start_row, 1).fill = PatternFill("solid", fgColor=BLUE)
    sheet.cell(start_row, 1).alignment = Alignment(vertical="center")
    table_start_row = start_row + 1

    for cell in cells:
        excel_cell = sheet.cell(
            table_start_row + int(cell["row_index"]),
            int(cell["col_index"]) + 1,
            display_cell_text(cell["text_value"], cell["footnote_marker"]),
        )
        excel_cell.border = THIN_BORDER
        excel_cell.alignment = Alignment(wrap_text=True, vertical="center")
        excel_cell.font = Font(name="Arial", size=9, bold=bool(cell["is_header"]), color=TEXT)
        if cell["is_header"]:
            excel_cell.fill = PatternFill("solid", fgColor=BLUE)

    validation_table = validation_table_from_cells(table, cells)
    specs = load_rule_specs_by_id(connection, checks) if checks else {}
    applied_roles: dict[tuple[int, int], str] = {}
    for check in checks:
        spec = specs.get(check["rule_id"])
        for highlight in highlight_cells_for(check, spec, validation_table.matrix):
            if highlight.row_index > max_row or highlight.col_index > max_col:
                continue
            key = (highlight.row_index, highlight.col_index)
            current_role = applied_roles.get(key)
            if current_role == "target" and highlight.role != "target":
                continue
            applied_roles[key] = highlight.role
            excel_cell = sheet.cell(table_start_row + highlight.row_index, highlight.col_index + 1)
            apply_highlight(excel_cell, status=check["status"], role=highlight.role)
            if highlight.role == "target":
                append_issue_comment(excel_cell, check)

    detail_start = table_start_row + max_row + 2
    sheet.cell(detail_start, 1, "검수 내용")
    sheet.cell(detail_start, 1).font = Font(name="Arial", size=11, bold=True, color=TEXT)
    detail_headers = ["상태", "검수 항목", "행", "열", "현재값", "검수값", "차이", "산식", "설명"]
    write_header_row(sheet, detail_start + 1, detail_headers)
    for offset, check in enumerate(checks, start=1):
        row_label, column_label = check_axis_labels(validation_table, check)
        values = [
            check["status"],
            check["check_type"],
            row_label,
            column_label,
            check["current_value"],
            check["expected_value"] or "",
            check["difference"] or "",
            check["formula"] or "",
            check["detail"] or "",
        ]
        write_data_row(sheet, detail_start + 1 + offset, values)
        status_cell = sheet.cell(detail_start + 1 + offset, 1)
        status_cell.fill = PatternFill(
            "solid",
            fgColor=ERROR_TARGET if check["status"] == "오류 의심" else REVIEW_TARGET,
        )
        status_cell.font = Font(name="Arial", size=9, bold=True, color=TEXT)
    if not checks:
        sheet.cell(detail_start + 2, 1, "이 표 부분에는 이상치 외 검수 문제가 없습니다.")
        sheet.cell(detail_start + 2, 1).font = Font(name="Arial", size=9, color=MUTED)
        return detail_start + 3
    return detail_start + 2 + len(checks)


def validation_table_from_cells(table: sqlite3.Row, cells: list[sqlite3.Row]) -> ValidationTable:
    from app.validation.models import ValidationCell

    return ValidationTable(
        id=int(table["id"]),
        report_id=int(table["report_id"]),
        code=table["code"],
        title=table["title"],
        unit=table["unit"],
        base_date=table["base_date"],
        source=table["source"],
        note=table["note"],
        cells=[
            ValidationCell(
                row_index=int(cell["row_index"]),
                col_index=int(cell["col_index"]),
                text_value=cell["text_value"],
                numeric_value=None,
                is_header=bool(cell["is_header"]),
            )
            for cell in cells
        ],
    )


def check_axis_labels(table: ValidationTable, check: sqlite3.Row) -> tuple[str, str]:
    row_index = check["row_index"]
    col_index = check["col_index"]
    if row_index is None or col_index is None:
        return check["location"] or "메타정보", "-"
    row_index = int(row_index)
    col_index = int(col_index)
    row_label = ""
    if 0 <= row_index < len(table.matrix):
        row_label = combined_row_label(table, table.matrix[row_index]) or table.row_label(table.matrix[row_index])
    column_label = table.column_text(col_index)
    return row_label or f"{row_index + 1}행", column_label or f"{col_index + 1}열"


def apply_highlight(cell, *, status: str, role: str) -> None:
    if role == "target":
        color = ERROR_TARGET if status == "오류 의심" else REVIEW_TARGET
        cell.font = Font(name="Arial", size=9, bold=True, color=TEXT)
        cell.border = Border(
            left=Side(style="medium", color="C00000" if status == "오류 의심" else "D97706"),
            right=Side(style="medium", color="C00000" if status == "오류 의심" else "D97706"),
            top=Side(style="medium", color="C00000" if status == "오류 의심" else "D97706"),
            bottom=Side(style="medium", color="C00000" if status == "오류 의심" else "D97706"),
        )
    else:
        color = ERROR_RELATED if status == "오류 의심" else REVIEW_RELATED
    cell.fill = PatternFill("solid", fgColor=color)


def append_issue_comment(cell, check: sqlite3.Row) -> None:
    body = (
        f"[{check['status']}] {check['check_type']}\n"
        f"현재값: {check['current_value']}\n"
        f"검수값: {check['expected_value'] or ''}\n"
        f"차이: {check['difference'] or ''}\n"
        f"{check['detail'] or ''}"
    )
    if cell.comment:
        body = f"{cell.comment.text}\n\n{body}"
    cell.comment = Comment(body[:32000], "통계연보 검수")


def write_validation_index_workbook(
    path: Path,
    *,
    checks: list[sqlite3.Row],
    contacts: dict[int, ContactMetadata],
    highlighted_filename: str,
    sheet_names: dict[str, str],
    validation_tables: dict[int, ValidationTable],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "검수내역"
    headers = ["번호", "통계명", "검수내용", "소속", "주무관", "내선번호", "출처"]
    write_header_row(sheet, 1, headers)
    for row_index, check in enumerate(checks, start=2):
        contact = contacts[int(check["table_id"])]
        table = validation_tables[int(check["table_id"])]
        row_label, column_label = check_axis_labels(table, check)
        detail = (
            f"{check['check_type']}\n"
            f"표 구분: {check['code']}\n"
            f"행: {row_label}\n"
            f"열: {column_label}\n"
            f"현재값: {check['current_value']}\n"
            f"검수값: {check['expected_value'] or ''}\n"
            f"차이: {check['difference'] or ''}"
        )
        values = [
            base_table_code(check["code"]),
            check["title"],
            detail,
            contact.department,
            contact.officer,
            contact.extension,
            contact.source_reference,
        ]
        write_data_row(sheet, row_index, values)
        base_code = base_table_code(check["code"])
        sheet_name = sheet_names[base_code]
        sheet.cell(row_index, 1).hyperlink = f"{highlighted_filename}#'{sheet_name}'!A1"
        sheet.cell(row_index, 1).style = "Hyperlink"
        if check["status"] == "오류 의심":
            sheet.cell(row_index, 3).fill = PatternFill("solid", fgColor=ERROR_RELATED)
        else:
            sheet.cell(row_index, 3).fill = PatternFill("solid", fgColor=REVIEW_RELATED)
    format_tabular_sheet(sheet, widths=[18, 38, 72, 24, 18, 20, 48])
    workbook.save(path)


def validation_table_map(db_path: Path, report_id: int) -> dict[int, ValidationTable]:
    repository = ValidationRepository(db_path)
    return {table.id: table for table in repository.load_tables(report_id)}


def display_cell_text(text: str, footnote_marker: str) -> str:
    restored = restore_hyphenated_line_breaks(text or "")
    restored = restored.replace("\r\n", "\n").replace("\r", "\n").strip()
    marker = (footnote_marker or "").strip()
    return f"{restored}{marker}" if marker and not restored.endswith(marker) else restored


def clean_multiline_text(value: str) -> str:
    restored = restore_hyphenated_line_breaks(value)
    return restored.replace("\\*", "*").replace("\r\n", "\n").replace("\r", "\n").strip()


def base_table_code(code: str) -> str:
    return PART_SUFFIX_RE.sub("", code).strip()


def group_tables_by_base(tables: Iterable[sqlite3.Row]) -> dict[str, list[sqlite3.Row]]:
    grouped: dict[str, list[sqlite3.Row]] = {}
    for table in tables:
        grouped.setdefault(base_table_code(table["code"]), []).append(table)
    return grouped


def join_distinct(values: Iterable[str], *, separator: str = " / ") -> str:
    return separator.join(unique(value for value in values if value))


def unique_sheet_name(value: str, used: set[str]) -> str:
    base = INVALID_SHEET_CHARS_RE.sub("_", value).strip(" '")[:31] or "통계"
    candidate = base
    suffix = 2
    while candidate in used:
        ending = f"_{suffix}"
        candidate = f"{base[:31 - len(ending)]}{ending}"
        suffix += 1
    used.add(candidate)
    return candidate


def table_max_col(connection: sqlite3.Connection, table_id: int) -> int:
    row = connection.execute(
        "SELECT COALESCE(MAX(col_index), 0) AS max_col FROM stat_table_cells WHERE table_id = ?",
        (table_id,),
    ).fetchone()
    return int(row["max_col"])


def write_header_row(sheet, row_index: int, headers: list[str]) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row_index, column_index, header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def write_data_row(sheet, row_index: int, values: list[object]) -> None:
    for column_index, value in enumerate(values, start=1):
        cell = sheet.cell(row_index, column_index, value if value is not None else "")
        cell.font = Font(name="Arial", size=9, color=TEXT)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = THIN_BORDER


def format_tabular_sheet(sheet, *, widths: list[int]) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 28
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row].height = min(max(30, 15 * max_newlines(sheet, row)), 120)


def max_newlines(sheet, row_index: int) -> int:
    return max(
        (str(sheet.cell(row_index, column).value or "").count("\n") + 1 for column in range(1, sheet.max_column + 1)),
        default=1,
    )


def validate_generated_workbooks(
    metadata_path: Path,
    highlighted_path: Path,
    validation_index_path: Path,
    *,
    expected_metadata_rows: int,
    expected_validation_rows: int,
    expected_validation_sheets: int,
) -> None:
    metadata = load_workbook(metadata_path, read_only=True)
    highlighted = load_workbook(highlighted_path, read_only=True)
    validation_index = load_workbook(validation_index_path, read_only=True)
    try:
        if metadata["메타데이터"].max_row - 1 != expected_metadata_rows:
            raise RuntimeError("메타데이터 XLSX 행 수가 원본 통계표 수와 다릅니다.")
        if validation_index["검수내역"].max_row - 1 != expected_validation_rows:
            raise RuntimeError("검수내역 XLSX 행 수가 이상치 제외 검수 건수와 다릅니다.")
        if len(highlighted.sheetnames) != expected_validation_sheets:
            raise RuntimeError("하이라이트 XLSX 시트 수가 문제 통계 수와 다릅니다.")
    finally:
        metadata.close()
        highlighted.close()
        validation_index.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="통계연보 메타데이터와 검수 결과를 XLSX로 내보냅니다.")
    parser.add_argument("--db", type=Path, default=DB_PATH)
    parser.add_argument("--year", type=int, default=None)
    parser.add_argument("--report-id", type=int, default=None)
    parser.add_argument("--run-id", type=int, default=None)
    parser.add_argument("--output-dir", type=Path, default=Path("exports"))
    return parser


def main() -> None:
    args = build_parser().parse_args()
    result = export_validation_workbooks(
        args.db,
        year=args.year,
        report_id=args.report_id,
        run_id=args.run_id,
        output_dir=args.output_dir,
    )
    print(
        f"Exported report {result.report_id}, run {result.run_id}: "
        f"metadata {result.metadata_rows} rows, validation {result.validation_rows} rows, "
        f"highlighted {result.validation_sheets} sheets\n"
        f"- {result.metadata_path}\n"
        f"- {result.highlighted_path}\n"
        f"- {result.validation_index_path}"
    )


if __name__ == "__main__":
    main()
