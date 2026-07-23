from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.db.connection import DB_PATH, connect
from app.export.metadata_owner_audit import (
    LogicalTable,
    OnnaraUser,
    ReferenceMatch,
    ReferenceOwner,
    clean_text,
    compare_department,
    compare_officers,
    index_users_by_name,
    join_unique,
    latest_report_id,
    load_onnara_users,
    load_logical_tables,
    load_reference_owners,
    match_tables_to_references,
    phone_digits,
    resolve_expected_users,
)


NAVY = "17324D"
WHITE = "FFFFFF"
TEXT = "172033"
MUTED = "5F6F86"
RED = "F4CCCC"
AMBER = "FFF2CC"
GRAY = "F3F5F8"
THIN = Side(style="thin", color="D6DFEA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MISMATCH_HEADERS = [
    "통계번호",
    "통계명",
    "반영 필요 항목",
    "엑셀 기준 부서",
    "2026 연보 초안 부서",
    "엑셀 기준 주무관",
    "2026 연보 초안 주무관",
    "엑셀 기준 내선번호",
    "2026 연보 초안 내선번호",
]


@dataclass(frozen=True)
class OwnerMismatchRow:
    code: str
    title: str
    mismatch_fields: str
    expected_department: str
    actual_department: str
    expected_officers: str
    actual_officers: str
    expected_phones: str
    actual_phones: str
    match_method: str
    match_score: float


@dataclass(frozen=True)
class OwnerMismatchExportResult:
    output_path: Path
    report_id: int
    total_tables: int
    mismatch_rows: int
    skipped_reference_missing: int
    low_confidence_matches: int
    mismatch_counts: dict[str, int]


def export_owner_reflection_mismatches(
    *,
    reference_xlsx: Path,
    onnara_xls: Path | None = None,
    db_path: Path = DB_PATH,
    output_path: Path = Path("exports/2026_통계연보_담당자_반영_미완료.xlsx"),
    report_2026_id: int | None = None,
) -> OwnerMismatchExportResult:
    references = load_reference_owners(reference_xlsx)[2026]
    users_by_name = index_users_by_name(load_onnara_users(onnara_xls)) if onnara_xls else {}
    with connect(db_path) as connection:
        report_id = report_2026_id or latest_report_id(connection, 2026)
        tables = load_logical_tables(connection, report_id)

    matches = match_tables_to_references(tables, references, 2026)
    rows, skipped_reference_missing, low_confidence = build_owner_mismatch_rows(
        tables=tables,
        matches=matches,
        users_by_name=users_by_name,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_mismatch_workbook(
        output_path,
        rows=rows,
        reference_xlsx=reference_xlsx,
        onnara_xls=onnara_xls,
        db_path=db_path,
        report_id=report_id,
        total_tables=len(tables),
        skipped_reference_missing=skipped_reference_missing,
        low_confidence_matches=low_confidence,
    )
    validate_mismatch_workbook(output_path, expected_rows=len(rows))

    counts = Counter(
        field.strip()
        for row in rows
        for field in row.mismatch_fields.split(",")
        if field.strip()
    )
    return OwnerMismatchExportResult(
        output_path=output_path.resolve(),
        report_id=report_id,
        total_tables=len(tables),
        mismatch_rows=len(rows),
        skipped_reference_missing=skipped_reference_missing,
        low_confidence_matches=low_confidence,
        mismatch_counts=dict(counts),
    )


def build_owner_mismatch_rows(
    *,
    tables: Sequence[LogicalTable],
    matches: dict[str, ReferenceMatch],
    users_by_name: dict[str, list[OnnaraUser]] | None = None,
) -> tuple[list[OwnerMismatchRow], int, int]:
    rows: list[OwnerMismatchRow] = []
    skipped_reference_missing = 0
    low_confidence = 0
    user_index = users_by_name or {}

    for table in tables:
        match = matches[table.code]
        reference = match.reference
        if reference is None:
            low_confidence += 1
            rows.append(
                OwnerMismatchRow(
                    code=table.code,
                    title=table.title,
                    mismatch_fields="통계 매칭 확인",
                    expected_department="",
                    actual_department=table.department,
                    expected_officers="",
                    actual_officers=" / ".join(table.officers),
                    expected_phones="",
                    actual_phones=" / ".join(table.phones),
                    match_method=match.method,
                    match_score=match.score,
                )
            )
            continue
        if match.score < 0.7:
            low_confidence += 1

        expected_phones = expected_phone_text(reference, user_index)
        fields = mismatch_fields(table, reference, expected_phones)
        if not fields:
            continue
        if not reference.department and not reference.officers:
            skipped_reference_missing += 1
            continue
        rows.append(
            OwnerMismatchRow(
                code=table.code,
                title=table.title,
                mismatch_fields=", ".join(fields),
                expected_department=reference.department,
                actual_department=table.department,
                expected_officers=reference.officer_raw,
                actual_officers=" / ".join(table.officers),
                expected_phones=expected_phones,
                actual_phones=" / ".join(table.phones),
                match_method=match.method,
                match_score=match.score,
            )
        )
    return rows, skipped_reference_missing, low_confidence


def mismatch_fields(
    table: LogicalTable,
    reference: ReferenceOwner,
    expected_phones: str = "",
) -> list[str]:
    fields: list[str] = []
    if reference.department and compare_department(reference.department, table.department) != "일치":
        fields.append("부서")
    if reference.officers and compare_officers(reference.officers, table.officers) != "일치":
        fields.append("담당자")
    if expected_phones and phones_differ(expected_phones, table.phones):
        fields.append("내선번호")
    return fields


def expected_phone_text(
    reference: ReferenceOwner,
    users_by_name: dict[str, list[OnnaraUser]],
) -> str:
    if not reference.officers or not users_by_name:
        return ""
    resolutions = resolve_expected_users(reference.officers, reference.department, users_by_name)
    phones = [
        resolution.primary.phone
        for resolution in resolutions
        if resolution.primary is not None and resolution.primary.phone
    ]
    if phones:
        return join_unique(phones)
    return "온나라 확인 필요"


def phones_differ(expected_phones: str, actual_phones: Sequence[str]) -> bool:
    expected_digits = {
        phone_digits(phone)
        for phone in expected_phones.split(" / ")
        if phone_digits(phone)
    }
    if not expected_digits:
        return False
    actual_digits = {phone_digits(phone) for phone in actual_phones if phone_digits(phone)}
    return not expected_digits <= actual_digits


def write_mismatch_workbook(
    path: Path,
    *,
    rows: Sequence[OwnerMismatchRow],
    reference_xlsx: Path,
    onnara_xls: Path | None,
    db_path: Path,
    report_id: int,
    total_tables: int,
    skipped_reference_missing: int,
    low_confidence_matches: int,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "반영 미완료"
    sheet.sheet_view.showGridLines = False

    write_title(sheet, "2026 통계연보 담당자 반영 미완료 목록")
    summary_rows = [
        ("생성 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("2026 연보 초안 report_id", report_id),
        ("2026 연보 초안 통계 수", total_tables),
        ("반영 미완료 통계 수", len(rows)),
        ("통계명 매칭 확인 필요", low_confidence_matches),
        ("기준 목록 부서·담당자 공란으로 제외", skipped_reference_missing),
        ("기준 엑셀", str(reference_xlsx)),
        ("온나라 사용자 목록", str(onnara_xls) if onnara_xls else ""),
        ("검증 DB", str(db_path)),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=3):
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, value)
        for column in range(1, 3):
            cell = sheet.cell(row_index, column)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="Arial", size=9, color=TEXT)
        sheet.cell(row_index, 1).fill = PatternFill("solid", fgColor=GRAY)
        sheet.cell(row_index, 1).font = Font(name="Arial", size=9, bold=True, color=MUTED)

    header_row = 13
    write_header(sheet, header_row, MISMATCH_HEADERS)
    for offset, row in enumerate(rows, start=1):
        write_row(
            sheet,
            header_row + offset,
            [
                row.code,
                row.title,
                row.mismatch_fields,
                row.expected_department,
                row.actual_department,
                row.expected_officers,
                row.actual_officers,
                row.expected_phones,
                row.actual_phones,
            ],
        )
        fill = RED if any(field in row.mismatch_fields for field in ("부서", "담당자")) else AMBER
        sheet.cell(header_row + offset, 3).fill = PatternFill("solid", fgColor=fill)

    widths = [16, 44, 18, 26, 26, 24, 24, 22, 22]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.column_dimensions["B"].width = 44
    sheet.freeze_panes = "A14"
    sheet.auto_filter.ref = f"A{header_row}:I{sheet.max_row}"
    for row_index in range(header_row + 1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 42

    workbook.save(path)


def write_title(sheet, value: str) -> None:
    sheet.merge_cells("A1:I1")
    cell = sheet["A1"]
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Arial", size=15, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32


def write_header(sheet, row_index: int, headers: Sequence[str]) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row_index, column_index, header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    sheet.row_dimensions[row_index].height = 30


def write_row(sheet, row_index: int, values: Sequence[object]) -> None:
    for column_index, value in enumerate(values, start=1):
        cell = sheet.cell(row_index, column_index, clean_text(value))
        cell.font = Font(name="Arial", size=9, color=TEXT)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER


def validate_mismatch_workbook(path: Path, *, expected_rows: int) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != ["반영 미완료"]:
            raise RuntimeError("담당자 반영 미완료 XLSX 시트 구성이 예상과 다릅니다.")
        if workbook["반영 미완료"].max_row - 13 != expected_rows:
            raise RuntimeError("담당자 반영 미완료 XLSX 행 수가 예상과 다릅니다.")
    finally:
        workbook.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="2026 목록과 2026 DB의 담당자 반영 미완료 항목만 추출합니다.")
    parser.add_argument("--reference-xlsx", required=True, type=Path)
    parser.add_argument("--onnara-xls", type=Path, default=None)
    parser.add_argument("--db", type=Path, default=DB_PATH)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("exports/2026_통계연보_담당자_반영_미완료.xlsx"),
    )
    parser.add_argument("--report-2026-id", type=int, default=None)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    result = export_owner_reflection_mismatches(
        reference_xlsx=args.reference_xlsx,
        onnara_xls=args.onnara_xls,
        db_path=args.db,
        output_path=args.output,
        report_2026_id=args.report_2026_id,
    )
    print(
        f"Exported {result.mismatch_rows} mismatch rows from {result.total_tables} tables: "
        f"{result.output_path}\n"
        f"Report ID: {result.report_id}\n"
        f"Mismatch counts: {result.mismatch_counts}"
    )


if __name__ == "__main__":
    main()
