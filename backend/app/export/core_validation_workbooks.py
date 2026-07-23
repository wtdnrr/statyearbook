from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
import sqlite3
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

from app.core.contact_metadata import ContactMetadata, parse_contact_metadata
from app.db.connection import DB_PATH, connect
from app.db.schema import init_db
from app.export.validation_workbooks import (
    ERROR_RELATED,
    REVIEW_RELATED,
    base_table_code,
    check_axis_labels,
    format_tabular_sheet,
    resolve_report,
    resolve_run_id,
    validation_table_map,
    write_data_row,
    write_header_row,
    write_highlighted_workbook,
)


CORE_CHECK_TYPES = (
    "합계 검수",
    "비율 검수",
    "메타정보 검수",
    "파란색 표기 확인",
)


@dataclass(frozen=True)
class CoreValidationExportResult:
    report_id: int
    run_id: int
    validation_rows: int
    validation_sheets: int
    highlighted_path: Path
    validation_index_path: Path


def export_core_validation_workbooks(
    db_path: Path = DB_PATH,
    *,
    year: int | None = None,
    report_id: int | None = None,
    run_id: int | None = None,
    output_dir: Path = Path("exports"),
) -> CoreValidationExportResult:
    output_dir.mkdir(parents=True, exist_ok=True)
    with connect(db_path) as connection:
        init_db(connection)
        report = resolve_report(connection, report_id=report_id, year=year)
        resolved_report_id = int(report["id"])
        resolved_run_id = resolve_run_id(
            connection,
            report_id=resolved_report_id,
            run_id=run_id,
        )
        tables = connection.execute(
            "SELECT * FROM stat_tables WHERE report_id = ? ORDER BY table_order, id",
            (resolved_report_id,),
        ).fetchall()
        checks = load_core_failing_checks(connection, resolved_run_id)
        contacts = {
            int(table["id"]): parse_contact_metadata(table["source"] or "")
            for table in tables
        }

        year_label = str(report["year"])
        highlighted_path = output_dir / f"{year_label}_통계연보_4대검수_하이라이트.xlsx"
        validation_index_path = output_dir / f"{year_label}_통계연보_4대검수_검수내역.xlsx"
        sheet_names = write_highlighted_workbook(
            highlighted_path,
            connection=connection,
            report_id=resolved_report_id,
            run_id=resolved_run_id,
            tables=tables,
            checks=checks,
            contacts=contacts,
        )
        write_core_validation_index(
            validation_index_path,
            checks=checks,
            contacts=contacts,
            highlighted_filename=highlighted_path.name,
            sheet_names=sheet_names,
            validation_tables=validation_table_map(db_path, resolved_report_id),
        )

    validate_core_workbooks(
        highlighted_path,
        validation_index_path,
        expected_rows=len(checks),
        expected_sheets=len(sheet_names),
    )
    return CoreValidationExportResult(
        report_id=resolved_report_id,
        run_id=resolved_run_id,
        validation_rows=len(checks),
        validation_sheets=len(sheet_names),
        highlighted_path=highlighted_path.resolve(),
        validation_index_path=validation_index_path.resolve(),
    )


def load_core_failing_checks(
    connection: sqlite3.Connection,
    run_id: int,
) -> list[sqlite3.Row]:
    placeholders = ", ".join("?" for _ in CORE_CHECK_TYPES)
    return connection.execute(
        f"""
        SELECT vc.*, st.code, st.title, st.table_order
        FROM validation_checks vc
        JOIN stat_tables st ON st.id = vc.table_id
        WHERE vc.run_id = ?
          AND vc.status != '정상'
          AND vc.check_type IN ({placeholders})
        ORDER BY st.table_order,
                 CASE vc.check_type
                     WHEN '합계 검수' THEN 0
                     WHEN '비율 검수' THEN 1
                     WHEN '메타정보 검수' THEN 2
                     WHEN '파란색 표기 확인' THEN 3
                     ELSE 4
                 END,
                 CASE vc.status WHEN '오류 의심' THEN 0 ELSE 1 END,
                 vc.row_index IS NULL, vc.row_index, vc.col_index, vc.id
        """,
        (run_id, *CORE_CHECK_TYPES),
    ).fetchall()


def write_core_validation_index(
    path: Path,
    *,
    checks: list[sqlite3.Row],
    contacts: dict[int, ContactMetadata],
    highlighted_filename: str,
    sheet_names: dict[str, str],
    validation_tables: dict[int, Any],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "검수내역"
    headers = [
        "번호",
        "통계명",
        "표 구분",
        "상태",
        "검수 항목",
        "행",
        "열",
        "현재값",
        "검수값",
        "차이",
        "산식",
        "설명",
        "소속",
        "이름",
        "내선번호",
        "출처",
    ]
    write_header_row(sheet, 1, headers)
    for row_index, check in enumerate(checks, start=2):
        table_id = int(check["table_id"])
        contact = contacts[table_id]
        table = validation_tables[table_id]
        row_label, column_label = check_axis_labels(table, check)
        values = [
            base_table_code(check["code"]),
            check["title"],
            check["code"],
            check["status"],
            check["check_type"],
            row_label,
            column_label,
            check["current_value"],
            check["expected_value"] or "",
            check["difference"] or "",
            check["formula"] or "",
            check["detail"] or "",
            contact.department,
            contact.officer,
            contact.extension,
            contact.source_reference,
        ]
        write_data_row(sheet, row_index, values)
        base_code = base_table_code(check["code"])
        sheet.cell(row_index, 1).hyperlink = (
            f"{highlighted_filename}#'{sheet_names[base_code]}'!A1"
        )
        sheet.cell(row_index, 1).style = "Hyperlink"
        status_color = ERROR_RELATED if check["status"] == "오류 의심" else REVIEW_RELATED
        sheet.cell(row_index, 4).fill = PatternFill("solid", fgColor=status_color)

    format_tabular_sheet(
        sheet,
        widths=[16, 34, 18, 13, 20, 34, 34, 18, 18, 16, 44, 72, 24, 18, 20, 44],
    )
    workbook.save(path)


def validate_core_workbooks(
    highlighted_path: Path,
    validation_index_path: Path,
    *,
    expected_rows: int,
    expected_sheets: int,
) -> None:
    highlighted = load_workbook(highlighted_path, read_only=True)
    validation_index = load_workbook(validation_index_path, read_only=True)
    try:
        if validation_index["검수내역"].max_row - 1 != expected_rows:
            raise RuntimeError("4대 검수 내역 행 수가 DB의 미통과 검수 건수와 다릅니다.")
        if len(highlighted.sheetnames) != expected_sheets:
            raise RuntimeError("4대 검수 하이라이트 시트 수가 문제 통계 수와 다릅니다.")
        actual_types = {
            str(row[0].value or "")
            for row in validation_index["검수내역"].iter_rows(
                min_row=2,
                min_col=5,
                max_col=5,
            )
        }
        if actual_types - set(CORE_CHECK_TYPES):
            raise RuntimeError("4대 검수 외의 항목이 엑셀에 포함되었습니다.")
    finally:
        highlighted.close()
        validation_index.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="4대 핵심 검수 문제를 XLSX 두 개로 내보냅니다.")
    parser.add_argument("--db", type=Path, default=DB_PATH)
    parser.add_argument("--year", type=int, default=None)
    parser.add_argument("--report-id", type=int, default=None)
    parser.add_argument("--run-id", type=int, default=None)
    parser.add_argument("--output-dir", type=Path, default=Path("exports"))
    return parser


def main() -> None:
    args = build_parser().parse_args()
    result = export_core_validation_workbooks(
        args.db,
        year=args.year,
        report_id=args.report_id,
        run_id=args.run_id,
        output_dir=args.output_dir,
    )
    print(
        f"Exported report {result.report_id}, run {result.run_id}: "
        f"{result.validation_rows} issues, {result.validation_sheets} sheets\n"
        f"- {result.validation_index_path}\n"
        f"- {result.highlighted_path}"
    )


if __name__ == "__main__":
    main()
