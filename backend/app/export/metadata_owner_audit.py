from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
import re
import sqlite3
import unicodedata
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import xlrd

from app.db.schema import DB_PATH, connect
from app.export.validation_workbooks import base_table_code, parse_contact_metadata


REFERENCE_SHEETS = {2025: "2025 통계연보 목록", 2026: "2026 통계연보 목록"}
REFERENCE_START_ROWS = {2025: 3, 2026: 4}
ROLE_WORDS = {
    "주무관",
    "사무관",
    "서기관",
    "연구사",
    "전문관",
    "과장",
    "팀장",
    "센터장",
    "담당자",
    "경위",
    "소방경",
    "소방위",
    "행정실무원",
}
ARROW_RE = re.compile(r"\s*(?:-{1,2}>|→|⇒)\s*")
PERSON_TOKEN_RE = re.compile(r"[가-힣]{2,6}")
NON_TITLE_RE = re.compile(r"[^0-9a-z가-힣]+")

# 목록의 묶음 제목이나 2026년에 변경된 명칭으로 인해 단순 제목 비교가 불가능한 표입니다.
REFERENCE_TITLE_OVERRIDES: dict[int, dict[str, str]] = {
    2026: {
        "2-1-4-1": "보조금24",
        "2-1-4-2": "보조금24",
        "2-1-4-3": "보조금24",
        "4-1-7-2": "연도별 공무원 정원",
        "4-1-11-2": "주민소환투표 청구 및 실시사례(2025년이후)",
        "4-2-12": "지역별 온천정보",
        "5-2-1-3": "지역별･세목별 지방세 규모",
        "7-2-1-3": "규모 4.0이상 지진 발생",
    },
}

NAVY = "17324D"
WHITE = "FFFFFF"
TEXT = "172033"
MUTED = "5F6F86"
BLUE = "DCEAF7"
GREEN = "E2F0D9"
AMBER = "FFF2CC"
RED = "F4CCCC"
GRAY = "F3F5F8"
THIN = Side(style="thin", color="D6DFEA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


@dataclass(frozen=True)
class ReferenceOwner:
    year: int
    row_number: int
    sequence_index: int
    title: str
    department_raw: str
    department: str
    officer_raw: str
    officers: tuple[str, ...]
    note: str


@dataclass(frozen=True)
class LogicalTable:
    report_id: int
    code: str
    title: str
    table_order: int
    department: str
    officers: tuple[str, ...]
    phones: tuple[str, ...]
    sources: tuple[str, ...]
    part_ids: tuple[int, ...]


@dataclass(frozen=True)
class ReferenceMatch:
    reference: ReferenceOwner | None
    method: str
    score: float


@dataclass(frozen=True)
class OnnaraUser:
    row_number: int
    display_department: str
    department: str
    name: str
    rank: str
    position: str
    phone: str


@dataclass(frozen=True)
class UserResolution:
    primary: OnnaraUser | None
    candidates: tuple[OnnaraUser, ...]
    status: str


@dataclass(frozen=True)
class AuditResult:
    output_path: Path
    total_rows: int
    action_rows: int
    changed_rows: int
    unmatched_db_rows: int
    unmatched_reference_rows: int
    decision_counts: dict[str, int]


@dataclass(frozen=True)
class AuditRow:
    values: dict[str, object]
    decision: str


AUDIT_HEADERS = [
    "번호",
    "통계명",
    "제목 매칭 방식",
    "제목 매칭 신뢰도",
    "2025 목록 부서",
    "2025 목록 담당자",
    "2025 DB 부서",
    "2025 DB 담당자",
    "2026 목록 부서(원문)",
    "2026 적용 부서",
    "2026 목록 담당자",
    "2026 DB 부서",
    "2026 DB 담당자",
    "2026 DB 번호",
    "온나라 부서",
    "온나라 성명",
    "온나라 번호",
    "부서 반영",
    "담당자 반영",
    "번호 검증",
    "온나라 기준 검증",
    "전년 대비 변경",
    "변경 반영 상태",
    "최종 판정",
    "조치사항",
    "목록 비고",
]


def export_metadata_owner_audit(
    *,
    reference_xlsx: Path,
    onnara_xls: Path,
    db_path: Path = DB_PATH,
    output_path: Path = Path("exports/2026_통계연보_담당자_메타데이터_검증.xlsx"),
    report_2025_id: int | None = None,
    report_2026_id: int | None = None,
) -> AuditResult:
    references = load_reference_owners(reference_xlsx)
    users = load_onnara_users(onnara_xls)
    with connect(db_path) as connection:
        report_2025_id = report_2025_id or latest_report_id(connection, 2025)
        report_2026_id = report_2026_id or latest_report_id(connection, 2026)
        tables_2025 = load_logical_tables(connection, report_2025_id)
        tables_2026 = load_logical_tables(connection, report_2026_id)

    matches_2025 = match_tables_to_references(tables_2025, references[2025], 2025)
    matches_2026 = match_tables_to_references(tables_2026, references[2026], 2026)
    prior_tables = match_prior_tables(tables_2026, tables_2025)
    users_by_name = index_users_by_name(users)

    rows = build_audit_rows(
        tables_2026=tables_2026,
        matches_2026=matches_2026,
        matches_2025=matches_2025,
        prior_tables=prior_tables,
        users_by_name=users_by_name,
    )
    matched_reference_rows = {
        match.reference.row_number
        for match in matches_2026.values()
        if match.reference is not None
    }
    unmatched_references = [
        reference
        for reference in references[2026]
        if reference.row_number not in matched_reference_rows
    ]
    unmatched_db = [
        table for table in tables_2026 if matches_2026[table.code].reference is None
    ]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    write_audit_workbook(
        output_path,
        rows=rows,
        unmatched_references=unmatched_references,
        unmatched_db=unmatched_db,
        reference_xlsx=reference_xlsx,
        onnara_xls=onnara_xls,
        db_path=db_path,
        report_2025_id=report_2025_id,
        report_2026_id=report_2026_id,
    )
    validate_audit_workbook(output_path, expected_rows=len(rows))

    decisions = Counter(row.decision for row in rows)
    return AuditResult(
        output_path=output_path.resolve(),
        total_rows=len(rows),
        action_rows=sum(row.decision != "정상" for row in rows),
        changed_rows=sum(is_actual_change(str(row.values["전년 대비 변경"])) for row in rows),
        unmatched_db_rows=len(unmatched_db),
        unmatched_reference_rows=len(unmatched_references),
        decision_counts=dict(decisions),
    )


def load_reference_owners(path: Path) -> dict[int, list[ReferenceOwner]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        result: dict[int, list[ReferenceOwner]] = {}
        for year, sheet_name in REFERENCE_SHEETS.items():
            if sheet_name not in workbook.sheetnames:
                raise RuntimeError(f"기준 파일에서 '{sheet_name}' 시트를 찾을 수 없습니다.")
            sheet = workbook[sheet_name]
            records: list[ReferenceOwner] = []
            for row_number, values in enumerate(
                sheet.iter_rows(
                    min_row=REFERENCE_START_ROWS[year],
                    max_col=5,
                    values_only=True,
                ),
                start=REFERENCE_START_ROWS[year],
            ):
                _, title, department, officer, note = values
                if not title or not (department or officer):
                    continue
                department_raw = clean_text(department)
                officer_raw = clean_text(officer)
                records.append(
                    ReferenceOwner(
                        year=year,
                        row_number=row_number,
                        sequence_index=len(records),
                        title=clean_text(title),
                        department_raw=department_raw,
                        department=current_department(department_raw),
                        officer_raw=officer_raw,
                        officers=parse_officer_names(officer_raw),
                        note=clean_text(note),
                    )
                )
            result[year] = records
        return result
    finally:
        workbook.close()


def load_onnara_users(path: Path) -> list[OnnaraUser]:
    workbook = xlrd.open_workbook(path, on_demand=True)
    try:
        sheet = workbook.sheet_by_index(0)
        expected_headers = ["부서표시명", "부서명", "성명", "직급", "직위", "사무실번호"]
        actual_headers = [clean_text(value) for value in sheet.row_values(0)[:6]]
        if actual_headers != expected_headers:
            raise RuntimeError(
                "온나라 사용자 목록의 열 구성이 예상과 다릅니다: " + ", ".join(actual_headers)
            )
        users: list[OnnaraUser] = []
        for row_index in range(1, sheet.nrows):
            values = [clean_text(value) for value in sheet.row_values(row_index)[:6]]
            if not values[2]:
                continue
            users.append(
                OnnaraUser(
                    row_number=row_index + 1,
                    display_department=values[0],
                    department=values[1],
                    name=values[2],
                    rank=values[3],
                    position=values[4],
                    phone=normalize_phone(values[5]),
                )
            )
        return users
    finally:
        workbook.release_resources()


def latest_report_id(connection: sqlite3.Connection, year: int) -> int:
    row = connection.execute(
        """
        SELECT id FROM annual_reports
        WHERE year = ?
        ORDER BY imported_at DESC, id DESC
        LIMIT 1
        """,
        (year,),
    ).fetchone()
    if row is None:
        raise RuntimeError(f"DB에서 {year}년 통계연보를 찾을 수 없습니다.")
    return int(row["id"])


def load_logical_tables(connection: sqlite3.Connection, report_id: int) -> list[LogicalTable]:
    physical_tables = connection.execute(
        "SELECT * FROM stat_tables WHERE report_id = ? ORDER BY table_order, id",
        (report_id,),
    ).fetchall()
    grouped: dict[str, list[sqlite3.Row]] = {}
    for table in physical_tables:
        grouped.setdefault(base_table_code(table["code"]), []).append(table)

    result: list[LogicalTable] = []
    for code, parts in grouped.items():
        contacts = [parse_contact_metadata(part["source"] or "") for part in parts]
        result.append(
            LogicalTable(
                report_id=report_id,
                code=code,
                title=clean_text(parts[0]["title"]),
                table_order=min(int(part["table_order"]) for part in parts),
                department=join_unique(contact.department for contact in contacts),
                officers=tuple(
                    unique(
                        name
                        for contact in contacts
                        for name in split_joined(contact.officer)
                    )
                ),
                phones=tuple(
                    unique(
                        normalize_phone(phone)
                        for contact in contacts
                        for phone in split_joined(contact.extension)
                    )
                ),
                sources=tuple(
                    unique(contact.source_reference for contact in contacts if contact.source_reference)
                ),
                part_ids=tuple(int(part["id"]) for part in parts),
            )
        )
    return result


def match_tables_to_references(
    tables: Sequence[LogicalTable],
    references: Sequence[ReferenceOwner],
    year: int,
) -> dict[str, ReferenceMatch]:
    exact: dict[str, list[ReferenceOwner]] = defaultdict(list)
    for reference in references:
        exact[normalize_title(reference.title)].append(reference)

    overrides = REFERENCE_TITLE_OVERRIDES.get(year, {})
    matches: dict[str, ReferenceMatch] = {}
    for table_index, table in enumerate(tables):
        target_title = overrides.get(table.code)
        if target_title:
            candidates = exact.get(normalize_title(target_title), [])
            selected = closest_by_position(candidates, table_index, len(tables), len(references))
            matches[table.code] = ReferenceMatch(selected, "번호별 명칭/묶음 매칭", 0.98 if selected else 0.0)
            continue

        candidates = exact.get(normalize_title(table.title), [])
        if candidates:
            selected = closest_by_position(candidates, table_index, len(tables), len(references))
            matches[table.code] = ReferenceMatch(selected, "제목 일치", 1.0)
            continue

        scored = sorted(
            (
                title_position_score(table, table_index, len(tables), reference, len(references))
                for reference in references
            ),
            key=lambda item: item[0],
            reverse=True,
        )
        if not scored:
            matches[table.code] = ReferenceMatch(None, "미매칭", 0.0)
            continue
        combined, title_score, reference = scored[0]
        if title_score < 0.58 or combined < 0.61:
            matches[table.code] = ReferenceMatch(None, "미매칭", round(combined, 3))
        else:
            method = "유사 제목 매칭" if combined >= 0.78 else "유사 제목 확인 필요"
            matches[table.code] = ReferenceMatch(reference, method, round(combined, 3))
    return matches


def closest_by_position(
    candidates: Sequence[ReferenceOwner],
    table_index: int,
    table_count: int,
    reference_count: int,
) -> ReferenceOwner | None:
    if not candidates:
        return None
    target = normalized_position(table_index, table_count)
    return min(
        candidates,
        key=lambda item: abs(normalized_position(item.sequence_index, reference_count) - target),
    )


def title_position_score(
    table: LogicalTable,
    table_index: int,
    table_count: int,
    reference: ReferenceOwner,
    reference_count: int,
) -> tuple[float, float, ReferenceOwner]:
    title_score = SequenceMatcher(
        None,
        normalize_title(table.title),
        normalize_title(reference.title),
    ).ratio()
    distance = abs(
        normalized_position(table_index, table_count)
        - normalized_position(reference.sequence_index, reference_count)
    )
    position_score = max(0.0, 1.0 - distance)
    return title_score * 0.9 + position_score * 0.1, title_score, reference


def match_prior_tables(
    current_tables: Sequence[LogicalTable],
    prior_tables: Sequence[LogicalTable],
) -> dict[str, LogicalTable | None]:
    exact: dict[str, list[LogicalTable]] = defaultdict(list)
    by_code = {table.code: table for table in prior_tables}
    for table in prior_tables:
        exact[normalize_title(table.title)].append(table)

    result: dict[str, LogicalTable | None] = {}
    for index, table in enumerate(current_tables):
        title_candidates = exact.get(normalize_title(table.title), [])
        if title_candidates:
            result[table.code] = min(
                title_candidates,
                key=lambda candidate: abs(candidate.table_order - table.table_order),
            )
            continue

        same_code = by_code.get(table.code)
        if same_code and title_similarity(table.title, same_code.title) >= 0.68:
            result[table.code] = same_code
            continue

        ranked = sorted(
            (
                (
                    title_similarity(table.title, candidate.title),
                    abs(
                        normalized_position(index, len(current_tables))
                        - normalized_position(prior_index, len(prior_tables))
                    ),
                    candidate,
                )
                for prior_index, candidate in enumerate(prior_tables)
            ),
            key=lambda item: (-item[0], item[1]),
        )
        result[table.code] = ranked[0][2] if ranked and ranked[0][0] >= 0.72 else None
    return result


def build_audit_rows(
    *,
    tables_2026: Sequence[LogicalTable],
    matches_2026: dict[str, ReferenceMatch],
    matches_2025: dict[str, ReferenceMatch],
    prior_tables: dict[str, LogicalTable | None],
    users_by_name: dict[str, list[OnnaraUser]],
) -> list[AuditRow]:
    rows: list[AuditRow] = []
    for table in tables_2026:
        match = matches_2026[table.code]
        reference = match.reference
        prior_table = prior_tables.get(table.code)
        prior_match = matches_2025.get(prior_table.code) if prior_table else None
        prior_reference = (
            prior_match.reference
            if prior_match and prior_match.reference is not None and prior_match.score >= 0.7
            else None
        )

        expected_department = reference.department if reference else ""
        expected_officers = reference.officers if reference else ()
        expected_resolution = resolve_expected_users(
            expected_officers,
            expected_department,
            users_by_name,
        )
        department_status = compare_department(expected_department, table.department)
        officer_status = compare_officers(expected_officers, table.officers)
        phone_status = compare_phone(
            expected_officers=expected_officers,
            db_officers=table.officers,
            db_phones=table.phones,
            resolutions=expected_resolution,
        )
        prior_department = prior_reference.department if prior_reference else ""
        prior_officers = prior_reference.officers if prior_reference else ()
        change_text = describe_change(
            prior_department,
            prior_officers,
            expected_department,
            expected_officers,
        )
        reflection = reflection_status(
            prior_department=prior_department,
            prior_officers=prior_officers,
            expected_department=expected_department,
            expected_officers=expected_officers,
            db_department=table.department,
            db_officers=table.officers,
        )
        onnara_status = summarize_user_resolutions(expected_resolution)
        decision = decide_status(
            match=match,
            reference=reference,
            department_status=department_status,
            officer_status=officer_status,
            phone_status=phone_status,
            user_status=onnara_status,
        )
        action = build_action(
            decision=decision,
            expected_department=expected_department,
            expected_officers=expected_officers,
            table=table,
            phone_status=phone_status,
            resolutions=expected_resolution,
            match=match,
        )
        onnara_users = [
            resolution.primary
            for resolution in expected_resolution
            if resolution.primary is not None
        ]
        values = {
            "번호": table.code,
            "통계명": table.title,
            "제목 매칭 방식": match.method,
            "제목 매칭 신뢰도": match.score,
            "2025 목록 부서": prior_reference.department_raw if prior_reference else "",
            "2025 목록 담당자": prior_reference.officer_raw if prior_reference else "",
            "2025 DB 부서": prior_table.department if prior_table else "",
            "2025 DB 담당자": " / ".join(prior_table.officers) if prior_table else "",
            "2026 목록 부서(원문)": reference.department_raw if reference else "",
            "2026 적용 부서": expected_department,
            "2026 목록 담당자": reference.officer_raw if reference else "",
            "2026 DB 부서": table.department,
            "2026 DB 담당자": " / ".join(table.officers),
            "2026 DB 번호": " / ".join(table.phones),
            "온나라 부서": join_unique(user.department for user in onnara_users),
            "온나라 성명": join_unique(user.name for user in onnara_users),
            "온나라 번호": join_unique(user.phone for user in onnara_users),
            "부서 반영": department_status,
            "담당자 반영": officer_status,
            "번호 검증": phone_status,
            "온나라 기준 검증": onnara_status,
            "전년 대비 변경": change_text,
            "변경 반영 상태": reflection,
            "최종 판정": decision,
            "조치사항": action,
            "목록 비고": reference.note if reference else "",
        }
        rows.append(AuditRow(values=values, decision=decision))
    return rows


def index_users_by_name(users: Iterable[OnnaraUser]) -> dict[str, list[OnnaraUser]]:
    result: dict[str, list[OnnaraUser]] = defaultdict(list)
    for user in users:
        result[normalize_name(user.name)].append(user)
    return result


def resolve_expected_users(
    names: Sequence[str],
    expected_department: str,
    users_by_name: dict[str, list[OnnaraUser]],
) -> list[UserResolution]:
    departments = split_expected_departments(expected_department)
    return [
        resolve_user(
            name,
            departments[index] if len(departments) == len(names) else expected_department,
            users_by_name.get(normalize_name(name), []),
        )
        for index, name in enumerate(names)
    ]


def resolve_user(
    name: str,
    expected_department: str,
    candidates: Sequence[OnnaraUser],
) -> UserResolution:
    candidates = tuple(candidates)
    if not candidates:
        return UserResolution(None, (), "성명 없음")

    department_candidates = tuple(
        candidate
        for candidate in candidates
        if department_matches(expected_department, candidate.department)
        or department_matches(expected_department, candidate.display_department)
    )
    if department_candidates:
        primary = sorted(department_candidates, key=lambda item: (not bool(item.phone), item.row_number))[0]
        suffix = f" (동명이인 {len(candidates)}명 중 부서 확인)" if len(candidates) > 1 else ""
        return UserResolution(primary, candidates, f"성명·부서 일치{suffix}")
    if len(candidates) == 1:
        return UserResolution(candidates[0], candidates, "성명 일치·부서 불일치")
    return UserResolution(None, candidates, f"동명이인 {len(candidates)}명·부서 확인 필요")


def compare_department(expected: str, actual: str) -> str:
    if not expected:
        return "2026 목록 부서 누락"
    if not actual:
        return "DB 부서 누락"
    return "일치" if department_matches(expected, actual) else "불일치"


def compare_officers(expected: Sequence[str], actual: Sequence[str]) -> str:
    if not expected:
        return "2026 목록 담당자 누락"
    if not actual:
        return "DB 담당자 누락"
    expected_names = {normalize_name(name) for name in expected}
    actual_names = {normalize_name(name) for name in actual}
    return "일치" if expected_names == actual_names else "불일치"


def compare_phone(
    *,
    expected_officers: Sequence[str],
    db_officers: Sequence[str],
    db_phones: Sequence[str],
    resolutions: Sequence[UserResolution],
) -> str:
    if not expected_officers:
        return "담당자 기준 없음"
    if {normalize_name(name) for name in expected_officers} != {
        normalize_name(name) for name in db_officers
    }:
        return "담당자 불일치로 비교 보류"
    expected_phones = {
        phone_digits(resolution.primary.phone)
        for resolution in resolutions
        if resolution.primary and resolution.primary.phone
    }
    if not expected_phones:
        return "온나라 번호 확인 불가"
    actual_phones = {phone_digits(phone) for phone in db_phones if phone}
    if not actual_phones:
        return "DB 번호 누락"
    return "일치" if expected_phones <= actual_phones else "불일치"


def summarize_user_resolutions(resolutions: Sequence[UserResolution]) -> str:
    if not resolutions:
        return "담당자 기준 없음"
    return " / ".join(unique(resolution.status for resolution in resolutions))


def describe_change(
    prior_department: str,
    prior_officers: Sequence[str],
    current_department: str,
    current_officers: Sequence[str],
) -> str:
    if not prior_department and not prior_officers:
        return "2025 매칭 없음"
    changes: list[str] = []
    if prior_department and current_department and not department_matches(prior_department, current_department):
        changes.append(f"부서 {prior_department} → {current_department}")
    if prior_officers and current_officers and {
        normalize_name(name) for name in prior_officers
    } != {normalize_name(name) for name in current_officers}:
        changes.append(f"담당자 {' / '.join(prior_officers)} → {' / '.join(current_officers)}")
    return "; ".join(changes) if changes else "변경 없음"


def reflection_status(
    *,
    prior_department: str,
    prior_officers: Sequence[str],
    expected_department: str,
    expected_officers: Sequence[str],
    db_department: str,
    db_officers: Sequence[str],
) -> str:
    expected_matches = (
        compare_department(expected_department, db_department) == "일치"
        and compare_officers(expected_officers, db_officers) == "일치"
    )
    changed = describe_change(
        prior_department,
        prior_officers,
        expected_department,
        expected_officers,
    )
    if changed == "2025 매칭 없음":
        return "2025 비교 불가" if not expected_matches else "2026 기준 일치"
    if changed == "변경 없음":
        return "변경 없음·일치" if expected_matches else "2026 기준 불일치"
    if expected_matches:
        return "변경 반영 완료"

    prior_department_matches = bool(prior_department) and department_matches(prior_department, db_department)
    prior_officer_matches = bool(prior_officers) and compare_officers(prior_officers, db_officers) == "일치"
    if prior_department_matches or prior_officer_matches:
        return "미반영(2025 값 잔존)"
    return "변경 후 값 불일치"


def decide_status(
    *,
    match: ReferenceMatch,
    reference: ReferenceOwner | None,
    department_status: str,
    officer_status: str,
    phone_status: str,
    user_status: str,
) -> str:
    if reference is None or match.score < 0.7:
        return "통계 매칭 확인 필요"
    if department_status in {"불일치", "DB 부서 누락"} or officer_status in {
        "불일치",
        "DB 담당자 누락",
    }:
        return "수정 필요"
    if department_status.startswith("2026 목록") or officer_status.startswith("2026 목록"):
        return "기준 목록 확인 필요"
    if (
        "불일치" in user_status
        or "없음" in user_status
        or user_status.startswith("동명이인")
        or " / 동명이인" in user_status
    ):
        return "온나라 확인 필요"
    if phone_status in {"불일치", "DB 번호 누락"}:
        return "번호 수정 필요"
    if phone_status == "온나라 번호 확인 불가":
        return "온나라 확인 필요"
    return "정상"


def build_action(
    *,
    decision: str,
    expected_department: str,
    expected_officers: Sequence[str],
    table: LogicalTable,
    phone_status: str,
    resolutions: Sequence[UserResolution],
    match: ReferenceMatch,
) -> str:
    actions: list[str] = []
    if decision == "통계 매칭 확인 필요":
        actions.append(f"통계명 매칭 수동 확인({match.method}, {match.score:.3f})")
    if expected_department and compare_department(expected_department, table.department) != "일치":
        actions.append(f"부서: '{table.department or '-'}' → '{expected_department}'")
    if expected_officers and compare_officers(expected_officers, table.officers) != "일치":
        actions.append(
            f"담당자: '{' / '.join(table.officers) or '-'}' → '{' / '.join(expected_officers)}'"
        )
    expected_phones = join_unique(
        resolution.primary.phone
        for resolution in resolutions
        if resolution.primary and resolution.primary.phone
    )
    if phone_status in {"불일치", "DB 번호 누락"} and expected_phones:
        actions.append(f"번호: '{' / '.join(table.phones) or '-'}' → '{expected_phones}'")
    elif phone_status == "온나라 번호 확인 불가":
        actions.append("온나라 목록에 사무실번호가 없어 DB 번호를 교차 검증할 수 없음")
    for name, resolution in zip(expected_officers, resolutions):
        if resolution.status == "성명 일치·부서 불일치" and resolution.primary:
            actions.append(
                f"온나라 확인: '{name}' 담당자는 '{resolution.primary.department}' 소속으로 조회됨"
            )
        elif resolution.primary is None:
            candidate_text = ", ".join(
                f"{candidate.name}/{candidate.department}/{candidate.phone or '-'}"
                for candidate in resolution.candidates[:5]
            )
            actions.append(
                f"온나라 확인: {name} {resolution.status}"
                + (f" ({candidate_text})" if candidate_text else "")
            )
    return "\n".join(unique(actions)) or "-"


def write_audit_workbook(
    path: Path,
    *,
    rows: Sequence[AuditRow],
    unmatched_references: Sequence[ReferenceOwner],
    unmatched_db: Sequence[LogicalTable],
    reference_xlsx: Path,
    onnara_xls: Path,
    db_path: Path,
    report_2025_id: int,
    report_2026_id: int,
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "요약"
    action_rows = [row for row in rows if row.decision != "정상"]
    changed_rows = [
        row for row in rows if is_actual_change(str(row.values["전년 대비 변경"]))
    ]

    summary_rows = [
        ("생성 일시", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("2026 DB 통계 수", len(rows)),
        ("수정·확인 필요", len(action_rows)),
        ("정상", len(rows) - len(action_rows)),
        ("전년 대비 부서·담당자 변경", len(changed_rows)),
        ("2026 DB 미매칭", len(unmatched_db)),
        ("2026 목록 미매칭", len(unmatched_references)),
        ("2025 DB report_id", report_2025_id),
        ("2026 DB report_id", report_2026_id),
        ("기준 목록", str(reference_xlsx)),
        ("온나라 사용자 목록", str(onnara_xls)),
        ("검증 DB", str(db_path)),
    ]
    write_title(summary, "2026 통계연보 담당자 메타데이터 검증")
    write_header(summary, 3, ["항목", "값"])
    for row_index, values in enumerate(summary_rows, start=4):
        write_row(summary, row_index, values)

    decision_start = 4 + len(summary_rows) + 2
    write_header(summary, decision_start, ["판정", "건수"])
    for offset, (decision, count) in enumerate(sorted(Counter(row.decision for row in rows).items()), 1):
        write_row(summary, decision_start + offset, [decision, count])
        apply_decision_fill(summary.cell(decision_start + offset, 1), decision)

    note_start = decision_start + len(Counter(row.decision for row in rows)) + 3
    summary.cell(note_start, 1, "판정 기준")
    summary.cell(note_start, 1).font = Font(name="Arial", size=11, bold=True, color=TEXT)
    notes = [
        "2026 목록의 부서·담당자를 우선 기준으로 DB 반영 여부를 확인했습니다.",
        "2025 목록과 달라진 항목이 DB에 2025 값으로 남으면 '미반영(2025 값 잔존)'으로 표시했습니다.",
        "온나라 목록은 동일 성명의 현재 부서와 사무실 번호를 교차 확인하는 근거로 사용했습니다.",
        "부서가 '기존부서->신부서' 형식이면 오른쪽 신부서를 2026 적용 부서로 사용했습니다.",
        "동명이인 또는 온나라 부서 불일치는 자동 확정하지 않고 확인 필요로 분류했습니다.",
    ]
    for offset, note in enumerate(notes, 1):
        summary.cell(note_start + offset, 1, f"• {note}")
        summary.cell(note_start + offset, 1).font = Font(name="Arial", size=10, color=TEXT)
    summary.column_dimensions["A"].width = 38
    summary.column_dimensions["B"].width = 95
    summary.sheet_view.showGridLines = False

    write_audit_sheet(workbook.create_sheet("수정·확인 필요"), action_rows)
    write_audit_sheet(workbook.create_sheet("전체 비교"), rows)
    write_audit_sheet(workbook.create_sheet("변경 목록"), changed_rows)
    write_unmatched_sheet(
        workbook.create_sheet("미매칭"),
        unmatched_references=unmatched_references,
        unmatched_db=unmatched_db,
    )
    workbook.save(path)


def write_audit_sheet(sheet, rows: Sequence[AuditRow]) -> None:
    write_header(sheet, 1, AUDIT_HEADERS)
    for row_index, audit_row in enumerate(rows, start=2):
        write_row(sheet, row_index, [audit_row.values[header] for header in AUDIT_HEADERS])
        apply_decision_fill(sheet.cell(row_index, AUDIT_HEADERS.index("최종 판정") + 1), audit_row.decision)
        if audit_row.values["변경 반영 상태"] == "미반영(2025 값 잔존)":
            sheet.cell(row_index, AUDIT_HEADERS.index("변경 반영 상태") + 1).fill = PatternFill(
                "solid", fgColor=RED
            )
    widths = [
        15, 36, 22, 15, 24, 18, 24, 18, 34, 24, 20, 24, 20, 20,
        30, 16, 20, 18, 18, 22, 30, 38, 24, 22, 72, 42,
    ]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    sheet.row_dimensions[1].height = 34
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 45


def write_unmatched_sheet(
    sheet,
    *,
    unmatched_references: Sequence[ReferenceOwner],
    unmatched_db: Sequence[LogicalTable],
) -> None:
    headers = ["구분", "번호/원본 행", "통계명", "부서", "담당자", "비고"]
    write_header(sheet, 1, headers)
    row_index = 2
    for reference in unmatched_references:
        write_row(
            sheet,
            row_index,
            [
                "2026 목록에만 존재",
                f"목록 {reference.row_number}행",
                reference.title,
                reference.department_raw,
                reference.officer_raw,
                reference.note,
            ],
        )
        row_index += 1
    for table in unmatched_db:
        write_row(
            sheet,
            row_index,
            [
                "2026 DB에만 존재",
                table.code,
                table.title,
                table.department,
                " / ".join(table.officers),
                "기준 목록 통계명 매칭 필요",
            ],
        )
        row_index += 1
    for index, width in enumerate([24, 20, 42, 32, 20, 50], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False


def write_title(sheet, value: str) -> None:
    sheet.merge_cells("A1:B1")
    cell = sheet["A1"]
    cell.value = value
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.font = Font(name="Arial", size=16, bold=True, color=WHITE)
    cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 32


def write_header(sheet, row_index: int, headers: Sequence[str]) -> None:
    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row_index, column_index, header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def write_row(sheet, row_index: int, values: Sequence[object]) -> None:
    for column_index, value in enumerate(values, start=1):
        cell = sheet.cell(row_index, column_index, value if value is not None else "")
        cell.font = Font(name="Arial", size=9, color=TEXT)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = BORDER


def apply_decision_fill(cell, decision: str) -> None:
    if decision == "정상":
        color = GREEN
    elif decision in {"수정 필요", "번호 수정 필요"}:
        color = RED
    else:
        color = AMBER
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(name="Arial", size=9, bold=True, color=TEXT)


def validate_audit_workbook(path: Path, *, expected_rows: int) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        expected_sheets = {"요약", "수정·확인 필요", "전체 비교", "변경 목록", "미매칭"}
        if set(workbook.sheetnames) != expected_sheets:
            raise RuntimeError("담당자 검증 XLSX 시트 구성이 예상과 다릅니다.")
        if workbook["전체 비교"].max_row - 1 != expected_rows:
            raise RuntimeError("담당자 검증 XLSX의 전체 비교 행 수가 DB 통계 수와 다릅니다.")
    finally:
        workbook.close()


def current_department(raw: str) -> str:
    if not raw:
        return ""
    parts = [part.strip(" ,;/?") for part in ARROW_RE.split(raw) if part.strip(" ,;/?")]
    return parts[-1] if parts else raw.strip()


def split_expected_departments(value: str) -> list[str]:
    return [part.strip() for part in re.split(r"\s*/\s*", value) if part.strip()]


def parse_officer_names(raw: str) -> tuple[str, ...]:
    if not raw:
        return ()
    cleaned = raw
    for role in sorted(ROLE_WORDS, key=len, reverse=True):
        cleaned = cleaned.replace(role, " ")
    names = [
        token
        for token in PERSON_TOKEN_RE.findall(cleaned)
        if token not in ROLE_WORDS and not token.endswith(("담당", "변경", "완료"))
    ]
    return tuple(unique(names))


def is_actual_change(value: str) -> bool:
    return value not in {"", "변경 없음", "2025 매칭 없음"}


def department_matches(expected: str, actual: str) -> bool:
    expected_normalized = normalize_department(expected)
    actual_normalized = normalize_department(actual)
    if not expected_normalized or not actual_normalized:
        return False
    return (
        expected_normalized == actual_normalized
        or actual_normalized.endswith(expected_normalized)
        or expected_normalized.endswith(actual_normalized)
    )


def normalize_department(value: str) -> str:
    normalized = re.sub(
        r"[^0-9a-z가-힣]",
        "",
        unicodedata.normalize("NFKC", value or "").lower(),
    )
    return re.sub(r"담당관실$", "담당관", normalized)


def normalize_title(value: str) -> str:
    return NON_TITLE_RE.sub("", unicodedata.normalize("NFKC", value or "").lower())


def normalize_name(value: str) -> str:
    return re.sub(r"[^가-힣a-z]", "", unicodedata.normalize("NFKC", value or "").lower())


def title_similarity(left: str, right: str) -> float:
    return SequenceMatcher(None, normalize_title(left), normalize_title(right)).ratio()


def normalized_position(index: int, count: int) -> float:
    return index / max(count - 1, 1)


def normalize_phone(value: str) -> str:
    digits = phone_digits(value)
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10 and digits.startswith("02"):
        return f"{digits[:2]}-{digits[2:6]}-{digits[6:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return clean_text(value)


def phone_digits(value: str) -> str:
    return re.sub(r"\D", "", value or "")


def clean_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value)).strip()


def split_joined(value: str) -> list[str]:
    return [part.strip() for part in value.split(" / ") if part.strip()]


def join_unique(values: Iterable[str]) -> str:
    return " / ".join(unique(value for value in values if value))


def unique(values: Iterable[str]) -> list[str]:
    return list(dict.fromkeys(value.strip() for value in values if value and value.strip()))


def append_text(current: str, addition: str) -> str:
    if not current or current == "-":
        return addition
    return f"{current}\n{addition}"


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="통계연보 담당자 메타데이터를 기준 목록·온나라와 비교합니다.")
    parser.add_argument("--reference-xlsx", required=True, type=Path)
    parser.add_argument("--onnara-xls", required=True, type=Path)
    parser.add_argument("--db", type=Path, default=DB_PATH)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("exports/2026_통계연보_담당자_메타데이터_검증.xlsx"),
    )
    parser.add_argument("--report-2025-id", type=int, default=None)
    parser.add_argument("--report-2026-id", type=int, default=None)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    result = export_metadata_owner_audit(
        reference_xlsx=args.reference_xlsx,
        onnara_xls=args.onnara_xls,
        db_path=args.db,
        output_path=args.output,
        report_2025_id=args.report_2025_id,
        report_2026_id=args.report_2026_id,
    )
    print(
        f"Exported {result.total_rows} rows ({result.action_rows} need action, "
        f"{result.changed_rows} changed): {result.output_path}\n"
        f"Decisions: {result.decision_counts}\n"
        f"Unmatched DB/reference: {result.unmatched_db_rows}/{result.unmatched_reference_rows}"
    )


if __name__ == "__main__":
    main()
