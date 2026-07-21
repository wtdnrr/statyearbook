from __future__ import annotations

from dataclasses import replace
from datetime import date, datetime
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from app.db.schema import DB_PATH
from app.ingest.hwpx_importer import (
    TABLE_CODE_RE,
    domain_from_code,
    extract_unit_and_base_date,
    guess_header_count,
    split_bilingual,
)
from app.ingest.repository import ImportedTable, ReportImportRepository


METADATA_LABELS = {
    "표번호": "code",
    "번호": "code",
    "통계번호": "code",
    "통계명": "title",
    "표제목": "title",
    "제목": "title",
    "영문명": "title_en",
    "영문제목": "title_en",
    "단위": "unit",
    "기준일": "base_date",
    "기준일자": "base_date",
    "주석": "note",
    "출처": "source",
    "분야": "domain",
}


def import_excel(
    source_path: Path,
    *,
    db_path: Path = DB_PATH,
    year: int,
    title: str | None = None,
    run_validation: bool = True,
) -> dict[str, int | str]:
    source_path = source_path.expanduser().resolve()
    tables = parse_excel(source_path)
    if not tables:
        raise ValueError("엑셀에서 2열 이상의 통계표를 찾지 못했습니다.")

    result = ReportImportRepository(db_path).replace_report(
        source_path=source_path,
        year=year,
        title=title or f"{year} 행정안전통계연보",
        tables=tables,
    )
    validation_issues = 0
    if run_validation:
        from app.validation.run_validations import run_validations

        validation = run_validations(db_path, report_id=result.report_id)
        validation_issues = int(validation["issues"])

    return {
        "db_path": str(db_path),
        "source_file": str(source_path),
        "report_id": result.report_id,
        "tables": result.table_count,
        "cells": result.cell_count,
        "validation_issues": validation_issues,
    }


def parse_excel(source_path: Path) -> list[ImportedTable]:
    workbook = load_workbook(source_path, data_only=True, read_only=False)
    parsed: list[ImportedTable] = []
    code_counts: dict[str, int] = {}
    table_order = 0
    try:
        for worksheet in workbook.worksheets:
            matrix = worksheet_matrix(worksheet)
            for section in split_table_sections(matrix):
                table = parse_table_section(
                    section,
                    sheet_name=worksheet.title,
                    table_order=table_order,
                )
                if table is None:
                    continue
                count = code_counts.get(table.code, 0) + 1
                code_counts[table.code] = count
                if count > 1:
                    table = replace(table, code=f"{table.code} 표{count}")
                parsed.append(table)
                table_order += 1
    finally:
        workbook.close()
    return parsed


def worksheet_matrix(worksheet: Worksheet) -> list[list[str]]:
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    values = [
        [display_cell_value(worksheet.cell(row=row, column=col)) for col in range(1, max_col + 1)]
        for row in range(1, max_row + 1)
    ]
    for merged in worksheet.merged_cells.ranges:
        value = values[merged.min_row - 1][merged.min_col - 1]
        for row in range(merged.min_row - 1, merged.max_row):
            for col in range(merged.min_col - 1, merged.max_col):
                values[row][col] = value
    return trim_matrix(values)


def display_cell_value(cell: Cell) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y. %m. %d.")
    if isinstance(value, date):
        return value.strftime("%Y. %m. %d.")
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)) and "%" in str(cell.number_format):
        decimals = excel_percentage_decimals(str(cell.number_format))
        return f"{float(value) * 100:.{decimals}f}%"
    if isinstance(value, int):
        return f"{value:,}" if "," in str(cell.number_format) else str(value)
    if isinstance(value, float):
        if value.is_integer():
            return f"{int(value):,}" if "," in str(cell.number_format) else str(int(value))
        return str(value)
    return re.sub(r"\s+", " ", str(value)).strip()


def excel_percentage_decimals(number_format: str) -> int:
    match = re.search(r"\.(0+)\s*%", number_format)
    return len(match.group(1)) if match else 0


def split_table_sections(matrix: list[list[str]]) -> list[list[list[str]]]:
    starts: list[int] = []
    for row_index, row in enumerate(matrix):
        leading_text = " ".join(row[: min(4, len(row))])
        if TABLE_CODE_RE.search(leading_text):
            starts.append(row_index)
    if not starts:
        return [matrix]
    return [
        matrix[start : starts[index + 1] if index + 1 < len(starts) else len(matrix)]
        for index, start in enumerate(starts)
    ]


def parse_table_section(
    section: list[list[str]],
    *,
    sheet_name: str,
    table_order: int,
) -> ImportedTable | None:
    if not section:
        return None
    metadata = metadata_from_rows(section)
    full_text = "\n".join(" ".join(row) for row in section)
    code_match = TABLE_CODE_RE.search(full_text)
    sheet_code_match = TABLE_CODE_RE.search(sheet_name)
    code = (
        metadata.get("code")
        or (code_match.group(0) if code_match else "")
        or (sheet_code_match.group(0) if sheet_code_match else "")
        or sheet_name.strip()
    )
    code = code or f"excel-{table_order + 1}"

    title_text = metadata.get("title") or inferred_title(section, code=code, sheet_name=sheet_name)
    title_ko, title_en = split_bilingual(title_text)
    title_ko = metadata.get("title") or title_ko or sheet_name
    title_en = metadata.get("title_en") or title_en

    data_start = find_data_start(section, code=code)
    matrix = trim_matrix(section[data_start:])
    if not matrix or max((len(row) for row in matrix), default=0) < 2:
        return None

    inferred_unit, inferred_base_date = extract_unit_and_base_date(full_text)
    return ImportedTable(
        code=code,
        title=title_ko,
        title_en=title_en,
        section_title=title_ko,
        section_title_en=title_en,
        domain=metadata.get("domain") or domain_from_code(code),
        unit=metadata.get("unit") or inferred_unit,
        base_date=metadata.get("base_date") or inferred_base_date,
        section_file=sheet_name,
        table_order=table_order,
        note=metadata.get("note", ""),
        source=metadata.get("source", ""),
        raw_context=full_text,
        matrix=matrix,
        header_count=guess_header_count(matrix),
    )


def metadata_from_rows(rows: list[list[str]]) -> dict[str, str]:
    metadata: dict[str, str] = {}
    for row in rows[:20]:
        nonempty = [value for value in row if value]
        if len(nonempty) < 2:
            continue
        label = normalize_label(nonempty[0])
        key = METADATA_LABELS.get(label)
        if key and key not in metadata:
            metadata[key] = nonempty[1].strip()
    return metadata


def inferred_title(rows: list[list[str]], *, code: str, sheet_name: str) -> str:
    for row in rows[:12]:
        nonempty = [value for value in row if value]
        if not nonempty:
            continue
        if normalize_label(nonempty[0]) in METADATA_LABELS:
            continue
        line = " ".join(nonempty)
        if code not in line:
            continue
        candidate = line.replace(code, "", 1).strip(" -_:：")
        if candidate:
            return candidate
    return sheet_name.replace(code, "", 1).strip(" -_:：") or sheet_name


def find_data_start(rows: list[list[str]], *, code: str) -> int:
    for row_index, row in enumerate(rows):
        first = normalize_label(next((value for value in row if value), ""))
        if first in {"표", "데이터", "원본표"}:
            return min(row_index + 1, len(rows))
    for row_index, row in enumerate(rows):
        nonempty = [value for value in row if value]
        if len(nonempty) < 2:
            continue
        if code in " ".join(row):
            continue
        if normalize_label(nonempty[0]) in METADATA_LABELS:
            continue
        return row_index
    return 0


def normalize_label(value: str) -> str:
    return re.sub(r"[\s:：]", "", value).lower()


def trim_matrix(matrix: list[list[str]]) -> list[list[str]]:
    rows = [list(row) for row in matrix]
    while rows and not any(value for value in rows[0]):
        rows.pop(0)
    while rows and not any(value for value in rows[-1]):
        rows.pop()
    if not rows:
        return []
    width = max((max((index for index, value in enumerate(row) if value), default=-1) + 1 for row in rows), default=0)
    return [[*row[:width], *([""] * max(width - len(row), 0))] for row in rows]
